import csv
import io
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from inventory.models import Product, RawMaterial, InventoryTransaction
from production.models import ProductionOrder


@login_required
def reports_index(request):
    return render(request, 'reports/index.html')


@login_required
def inventory_report(request):
    products = Product.objects.select_related('category', 'default_location').order_by('product_name')
    low_stock = products.filter(current_stock__lte=F('minimum_stock'))

    # Date range filter
    try:
        days = int(request.GET.get('days', 30))
        days = max(1, min(days, 365))
    except (ValueError, TypeError):
        days = 30
    since = timezone.now() - timedelta(days=days)
    transactions = (
        InventoryTransaction.objects
        .filter(date__gte=since, product__isnull=False)
        .select_related('product', 'warehouse_location', 'performed_by')
        .order_by('-date')
    )

    total_in = transactions.filter(transaction_type='IN').aggregate(t=Sum('quantity'))['t'] or 0
    total_out = transactions.filter(transaction_type__in=['OUT', 'PRODUCTION']).aggregate(t=Sum('quantity'))['t'] or 0

    return render(request, 'reports/inventory_report.html', {
        'products': products,
        'low_stock': low_stock,
        'transactions': transactions[:50],
        'total_in': total_in,
        'total_out': total_out,
        'days': days,
    })


@login_required
def production_report(request):
    try:
        days = int(request.GET.get('days', 30))
        days = max(1, min(days, 365))
    except (ValueError, TypeError):
        days = 30
    since = timezone.now() - timedelta(days=days)

    orders = ProductionOrder.objects.filter(created_at__gte=since).select_related('product', 'created_by')
    completed = orders.filter(status='completed')
    in_production = orders.filter(status='in_production')
    pending = orders.filter(status='pending')
    cancelled = orders.filter(status='cancelled')

    # Material usage
    material_usage = (
        InventoryTransaction.objects
        .filter(transaction_type='PRODUCTION', date__gte=since, raw_material__isnull=False)
        .values('raw_material__name', 'raw_material__unit')
        .annotate(total=Sum('quantity'))
        .order_by('-total')
    )

    return render(request, 'reports/production_report.html', {
        'orders': orders,
        'completed': completed,
        'in_production': in_production,
        'pending': pending,
        'cancelled': cancelled,
        'material_usage': material_usage,
        'days': days,
    })


@login_required
def export_inventory_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'
    writer = csv.writer(response)
    writer.writerow(['Product Name', 'SKU', 'Category', 'Unit', 'Current Stock', 'Min Stock', 'Status', 'Location'])
    for p in Product.objects.select_related('category', 'default_location').order_by('product_name'):
        writer.writerow([
            p.product_name, p.sku,
            p.category.name if p.category else '',
            p.get_unit_display(), p.current_stock, p.minimum_stock,
            'Low Stock' if p.is_low_stock else 'OK',
            str(p.default_location) if p.default_location else '',
        ])
    return response


@login_required
def export_inventory_excel(request):
    if not HAS_OPENPYXL:
        return HttpResponse('openpyxl not installed', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory Report'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    headers = ['Product Name', 'SKU', 'Category', 'Unit', 'Current Stock', 'Min Stock', 'Status', 'Location']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, p in enumerate(Product.objects.select_related('category', 'default_location').order_by('product_name'), 2):
        ws.append([
            p.product_name, p.sku,
            p.category.name if p.category else '',
            p.get_unit_display(), float(p.current_stock), float(p.minimum_stock),
            'Low Stock' if p.is_low_stock else 'OK',
            str(p.default_location) if p.default_location else '',
        ])
        if p.is_low_stock:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="inventory_report.xlsx"'
    return response


@login_required
def export_production_csv(request):
    try:
        days = int(request.GET.get('days', 30))
        days = max(1, min(days, 365))
    except (ValueError, TypeError):
        days = 30
    since = timezone.now() - timedelta(days=days)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="production_report.csv"'
    writer = csv.writer(response)
    writer.writerow(['Order #', 'Product', 'Quantity', 'Status', 'Created By', 'Start', 'End', 'Created At'])
    for o in ProductionOrder.objects.filter(created_at__gte=since).select_related('product', 'created_by'):
        writer.writerow([
            o.order_number, o.product.product_name, o.quantity,
            o.get_status_display(),
            str(o.created_by) if o.created_by else '',
            o.actual_start.strftime('%Y-%m-%d %H:%M') if o.actual_start else '',
            o.actual_end.strftime('%Y-%m-%d %H:%M') if o.actual_end else '',
            o.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    return response


@login_required
def export_production_excel(request):
    if not HAS_OPENPYXL:
        return HttpResponse('openpyxl not installed', status=500)

    try:
        days = int(request.GET.get('days', 30))
        days = max(1, min(days, 365))
    except (ValueError, TypeError):
        days = 30
    since = timezone.now() - timedelta(days=days)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Production Report'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    headers = ['Order #', 'Product', 'Quantity', 'Status', 'Created By', 'Start', 'End', 'Created At']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    status_colors = {
        'completed': 'C8E6C9',
        'in_production': 'FFF9C4',
        'pending': 'E3F2FD',
        'cancelled': 'FFCDD2',
    }
    for o in ProductionOrder.objects.filter(created_at__gte=since).select_related('product', 'created_by'):
        row = [
            o.order_number, o.product.product_name, float(o.quantity),
            o.get_status_display(),
            str(o.created_by) if o.created_by else '',
            o.actual_start.strftime('%Y-%m-%d %H:%M') if o.actual_start else '',
            o.actual_end.strftime('%Y-%m-%d %H:%M') if o.actual_end else '',
            o.created_at.strftime('%Y-%m-%d %H:%M'),
        ]
        ws.append(row)
        color = status_colors.get(o.status, 'FFFFFF')
        for col in range(1, 9):
            ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="production_report.xlsx"'
    return response
